import smtplib
from openpyxl import load_workbook

server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
server.login("email id", "password") #email id and password

workbook = load_workbook(filename=r"C:\Users\MDD\Desktop\python101\data.xlsx") #filepath
sheet = workbook.active


message=''
for value in sheet.iter_rows(min_row=2,min_col=1,values_only=True):
	print(value)
	message=message + str(value)

server.sendmail("email id","receipient email id",message)