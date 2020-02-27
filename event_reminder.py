import smtplib,re,operator,sys
from openpyxl import load_workbook
from datetime import datetime,date
from io import StringIO


workbook = load_workbook(filename=r"C:\Users\MDD\Desktop\python101\data_1.xlsx") #filepath


sheet = workbook.active

birthday_count=0
anniversary_count=0
kids_birthdays_count=0
birthdays={}
anniversaries={}
kids_birthdays={}
this_weeks_birthday_count = 0
this_weeks_anniversary_count=0
this_weeks_kids_birthdays_count=0
this_weeks_birthdays={}
this_weeks_anniversaries={}
this_weeks_kids_birthdays={}
next_weeks_anniversary_count=0
next_weeks_birthday_count = 0
next_weeks_kids_birthdays_count=0
next_weeks_birthdays={}
next_weeks_anniversaries={}
next_weeks_kids_birthdays={}
next_month_birthday_count = 0
next_month_anniversary_count=0
next_month_kids_birthdays_count=0
next_month_birthdays={}
next_month_anniversaries={}
next_month_kids_birthdays={}


sender_email = "a@gmail.com"  # Enter your address
receiver_email = "d@gmail.com"  # Enter receiver address


old_stdout = sys.stdout # Memorize the default stdout stream
result= StringIO()

sys.stdout = result

def change_date_format(dt):
        return re.sub(r'(\d{4})-(\d{1,2})-(\d{1,2})', '\\3-\\2-\\1', dt)

def is_event_this_week(date_today,date_event):
	day_since= (date_event-date_today).days
	day_today=date_today.weekday()
	map_this_week={0:[0,6],1:[-1,5],2:[-2,4],3:[-3,3],4:[-4,2],5:[-5,1],6:[-6,0]}
	map_next_week={0:[7,13],1:[6,12],2:[5,11],3:[4,10],4:[3,9],5:[2,8],6:[1,7]}
	this_week_range=map_this_week[day_today]
	next_week_range=map_next_week[day_today]
	#print(day_today,date_today,day_since,date_event)
	if(day_since>=this_week_range[0] and day_since<=this_week_range[1]):
		return 'ThisWeek'
	else:
		if(day_since>=next_week_range[0] and day_since<=next_week_range[1]):
			return 'NextWeek'
		else:
			return False
	

for value in sheet.iter_rows(min_row=2,min_col=1,values_only=True):	
	if(isinstance(value[2],datetime)): #to check if birthdays column has values
		if(value[2].date().month==date.today().month): 
			event_this_week_or_not=is_event_this_week(date.today(),value[2].date())
			if event_this_week_or_not=='ThisWeek':
				this_weeks_birthday_count+=1
				this_weeks_birthdays.update({value[0]:value[2]})
			birthday_count+=1
			birthdays.update({value[0]:value[2]})
		if(date.today().month==12): #handling december month
			if(value[2].date().month==1):
				next_month_birthday_count+=1
				next_month_birthdays.update({value[0]:value[2]})
		else:
			#print("here")
			if(value[2].date().month==date.today().month+1):
				event_this_week_or_not=is_event_this_week(date.today(),value[2].date())
				if event_this_week_or_not=='NextWeek':
					next_weeks_birthday_count+=1
					next_weeks_birthdays.update({value[0]:value[2]})
				next_month_birthday_count+=1
				next_month_birthdays.update({value[0]:value[2]})


	if(isinstance(value[3],datetime)): #to check if anniverary column has values
		if(value[3].date().month==date.today().month):
			event_this_week_or_not=is_event_this_week(date.today(),value[3].date())
			if event_this_week_or_not:
				this_weeks_anniversary_count+=1
				this_weeks_anniversaries.update({value[0]:value[3]})
			anniversary_count+=1
			anniversaries.update({value[0]:value[3]})

		if(date.today().month==12):
			if(value[3].date().month==1):
				next_month_birthday_count+=1
				next_month_birthdays.update({value[0]:value[3]})
		else:
			if(value[3].date().month==date.today().month+1):
				next_month_birthday_count+=1
				next_month_birthdays.update({value[0]:value[3]})

	if(isinstance(value[4],datetime)): #to check if kids birthdays column has values
		if(value[4].date().month==date.today().month):
			event_this_week_or_not=is_event_this_week(date.today(),value[4].date())
			if event_this_week_or_not:
				this_weeks_kids_birthdays_count+=1
				this_weeks_kids_birthdays.update({value[0]:value[4]})
			kids_birthdays_count+=1
			kids_birthdays.update({value[0]:value[4]})

		if(date.today().month==12):
			if(value[4].date().month==1):
				next_month_birthday_count+=1
				next_month_birthdays.update({value[0]:value[4]})
		else:
			#print("here")
			if(value[4].date().month==date.today().month+1):
				next_month_birthday_count+=1
				next_month_birthdays.update({value[0]:value[4]})

#print(f"This Month has {birthday_count} birthdays, {anniversary_count} anniversaries & {kids_birthdays_count} kids birthdays \n")
this_weeks_birthdays=dict(sorted(this_weeks_birthdays.items(), key=operator.itemgetter(1)))
this_weeks_anniversaries=dict(sorted(this_weeks_anniversaries.items(), key=operator.itemgetter(1)))
this_weeks_kids_birthdays=dict(sorted(this_weeks_kids_birthdays.items(), key=operator.itemgetter(1)))

print("---------------------------------------------------")
if this_weeks_birthday_count:
	if this_weeks_anniversary_count:
		if this_weeks_kids_birthdays_count:
			print(f"This week has {this_weeks_birthday_count} birthdays and {this_weeks_anniversary_count} anniversaries and {this_weeks_kids_birthdays_count} kids birthdays \n")
		else:
			print(f"This week has {this_weeks_birthday_count} birthdays and {this_weeks_anniversary_count} anniversaries \n")
	else:
		if this_weeks_kids_birthdays_count:
			print(f"This week has {this_weeks_birthday_count} birthdays and {this_weeks_kids_birthdays_count} kids birthdays \n")
		else:
			print(f"This week has {this_weeks_birthday_count} birthdays \n")
else:
	if this_weeks_anniversary_count:
		if this_weeks_kids_birthdays:
			print(f"This week has {this_weeks_anniversary_count} anniversaries and {this_weeks_kids_birthdays_count} kids birthdays \n")
		else: 
			print(f"This week has {this_weeks_anniversary_count} anniversaries \n")
	else:
		if this_weeks_kids_birthdays:
			print(f"This week has {this_weeks_kids_birthdays_count} kids birthdays \n")


for key in this_weeks_birthdays:
	print("Birthday: "+key,'->',change_date_format(str(this_weeks_birthdays[key].date())))
	b_message = str(key) + "->" 
for key in this_weeks_anniversaries:
	print("Anniversary: "+key,'->',change_date_format(str(this_weeks_anniversaries[key].date())))
	b_message = str(key) + "->" 
for key in this_weeks_kids_birthdays:
	print("Kids Birthday: "+key,'->',change_date_format(str(this_weeks_kids_birthdays[key].date())))
	b_message = str(key) + "->" 

print("---------------------------------------------------")
print("\n")



#print(next_weeks_birthday_count,next_weeks_birthdays)

if next_weeks_birthday_count:
	if next_weeks_anniversary_count:
		if next_weeks_kids_birthdays_count:
			print(f"Next week has {next_weeks_birthday_count} birthdays and {next_weeks_anniversary_count} anniversaries and {next_weeks_kids_birthdays_count} kids birthdays \n")
		else:
			print(f"Next week has {next_weeks_birthday_count} birthdays and {next_weeks_anniversary_count} anniversaries \n")
	else:
		if next_weeks_kids_birthdays_count:
			print(f"Next week has {next_weeks_birthday_count} birthdays and {next_weeks_kids_birthdays_count} kids birthdays \n")
		else:
			print(f"Next week has {next_weeks_birthday_count} birthdays \n")
else:
	if next_weeks_anniversary_count:
		if next_weeks_kids_birthdays_count:
			print(f"Next week has {next_weeks_anniversary_count} anniversaries and {next_weeks_kids_birthdays_count} kids birthdays \n")
		else: 
			print(f"Next week has {next_weeks_anniversary_count} anniversaries \n")
	else:
		if next_weeks_kids_birthdays:
			print(f"Next week has {next_weeks_kids_birthdays_count} kids birthdays \n")

for key in next_weeks_birthdays:
	print("Birthday: "+key,'->',change_date_format(str(next_weeks_birthdays[key].date())))
	b_message = str(key) + "->" 
for key in next_weeks_anniversaries:
	print("Anniversary: "+key,'->',change_date_format(str(next_weeks_anniversaries[key].date())))
	b_message = str(key) + "->" 
for key in next_weeks_kids_birthdays:
	print("Kids Birthday: "+key,'->',change_date_format(str(next_weeks_kids_birthdays[key].date())))
	b_message = str(key) + "->" 

print("---------------------------------------------------")
print("\n")

birthdays=dict(sorted(birthdays.items(), key=operator.itemgetter(1)))
if birthday_count:
	print(f"This month has {birthday_count} birthdays: \n")
	for key in birthdays:
		print(key,'->',change_date_format(str(birthdays[key].date())))
		b_message = str(key) + "->" 
	print("\n")

anniversaries=dict(sorted(anniversaries.items(), key=operator.itemgetter(1)))
if anniversary_count:
	print(f"This month has {anniversary_count} anniversaries: \n")
	for key in anniversaries:
		print(key,'->',change_date_format(str(anniversaries[key].date())))	
	print("\n")

kids_birthdays=dict(sorted(kids_birthdays.items(), key=operator.itemgetter(1)))
if kids_birthdays_count:
	print(f"This month has {kids_birthdays_count} kids birthdays: \n")
	for key in kids_birthdays:
		print(key,'->',change_date_format(str(kids_birthdays[key].date())))
	print("\n")

print("---------------------------------------------------")
print(f"Next Month has {next_month_birthday_count} birthdays, {next_month_anniversary_count} anniversaries & {next_month_kids_birthdays_count} kids birthdays \n")

next_month_birthdays=dict(sorted(next_month_birthdays.items(), key=operator.itemgetter(1)))
print(f"Next month has {next_month_birthday_count} birthdays: \n")
for key in next_month_birthdays:
	print(key,'->',change_date_format(str(next_month_birthdays[key].date())))
	b_message = str(key) + "->" 
print("\n")

next_month_anniversaries=dict(sorted(next_month_anniversaries.items(), key=operator.itemgetter(1)))
print(f"Next month has {next_month_anniversary_count} anniversaries: \n")
for key in next_month_anniversaries:
	print(key,'->',change_date_format(str(next_month_anniversaries[key].date())))
print("\n")

next_month_kids_birthdays=dict(sorted(next_month_kids_birthdays.items(), key=operator.itemgetter(1)))
print(f"Next month has {next_month_kids_birthdays_count} kids birthdays: \n")
for key in next_month_kids_birthdays:
	print(key,'->',change_date_format(str(next_month_kids_birthdays[key].date())))
print("\n")
print("---------------------------------------------------")
total_events = birthday_count + anniversary_count + kids_birthdays_count
next_month_total_events= next_month_birthday_count + next_month_anniversary_count + next_month_kids_birthdays_count

sys.stdout = old_stdout # Put the old stream back in place
this_weeks_total_events = this_weeks_birthday_count+this_weeks_anniversary_count+this_weeks_kids_birthdays_count
#next_weeks_total_events = next_weeks_kids_birthdays_count+next_weeks_anniversary_count+next_weeks_kids_birthdays_count

if this_weeks_total_events:
	message = """\
	Subject: This Week's Events (Birthdays, Anniversaries etc): """ + str(this_weeks_total_events) + """
	"""
	whatWasPrinted = message + result.getvalue() # Return a str containing the entire contents of the buffer.

	print(f"message is {whatWasPrinted}")

	server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
	server.login("email id", "password") #email id and password
	server.sendmail(sender_email,receiver_email,whatWasPrinted)
else:
	print("No events this week")