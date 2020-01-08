from openpyxl import load_workbook

workbook = load_workbook(filename=r"C:\Users\Desktop\python101\income_level.xlsx")

sheet = workbook.active

name=input("Please enter your name: ")
monthly_income=input("Please enter your monthly income: ")
if "," in monthly_income:
	monthly_income=monthly_income.replace(",","")
slab=0

for value in sheet.iter_rows(min_row=2,min_col=1,values_only=True):
	if int(monthly_income)<int(value[0]):
		slab=int(value[0])
		#print(slab)
		if slab not in (1656,3840626):
			print(name + ", you belong to the " +str(value[3])+" ("+str(value[1])+"% to "+str(value[2])+"%) of India")
		elif slab==1656:
			print(name + ", you belong to the " +str(value[3])+" "+str(value[1])+"% of India")
		else:
			print(name + ", you belong to the " +str(value[3])+" "+str(value[1])+"% of India")	
		exit()
#except:
#	print("Please check the values you entered")
#try: